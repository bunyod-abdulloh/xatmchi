import os

from aiogram import types
from openpyxl.workbook import Workbook

from loader import db


async def statistika_func(msg):
    all_ = await db.select_all_xatms()

    cnt = 0
    cnt_half = 0

    wb = Workbook()

    ws = wb.active
    ws["A1"] = "Xatm ID si"
    ws["B1"] = "Qolgan poralar"
    ws["C1"] = "To'liq ma'lumot"
    for n in all_:
        a = str(cnt + cnt_half + 2)
        ws["A" + a] = n[0]
        ws["B" + a] = n[1]
        ws["C" + a] = n[2]

        if n[1] == "[]":
            cnt += 1
        else:
            cnt_half += 1

    wb.save("joriyholat_xatm.xlsx")

    await msg.answer_document(types.InputFile(path_or_bytesio="joriyholat_xatm.xlsx"),
                              caption="Жорий ҳолатни кўрсатувчи жадвал")
    xatmga_qatnashgan_users = await db.select_xatmdagi_user(poralar="{}")
    users = len(xatmga_qatnashgan_users)
    await msg.answer(f"Жами очилган хатмлар: {cnt + cnt_half}"
                     f"\nТўлиқ банд қилинган хатмлар: {cnt}"
                     f"\nЖами қатнашчилар: {users}")
    os.remove("joriyholat_xatm.xlsx")


async def sarhisob_func(msg):
    all_ = await db.select_all_xatms()

    cnt = 0
    cnt_half = 0

    wb = Workbook()

    ws = wb.active
    ws["A1"] = "Xatm ID si"
    ws["B1"] = "Qolgan poralar"
    ws["C1"] = "To'liq ma'lumot"
    for n in all_:
        a = str(cnt + cnt_half + 2)
        ws["A" + a] = n[0]
        ws["B" + a] = n[1]
        ws["C" + a] = n[2]

        if n[1] == "[]":
            cnt += 1
        else:
            cnt_half += 1

    wb.save("sarhisob_xatm.xlsx")

    await msg.answer_document(types.InputFile(path_or_bytesio="sarhisob_xatm.xlsx"))
    xatmga_qatnashgan_users = await db.select_xatmdagi_user(poralar="{}")
    users = len(xatmga_qatnashgan_users)
    await msg.answer(f"Жами очилган хатмлар: {cnt + cnt_half}"
                     f"\nЯкунланган хатмлар: {cnt}"
                     f"\nЖами қатнашчилар: {users}")
    os.remove("sarhisob_xatm.xlsx")


async def user_xatmonalar(msg):
    all_stat = await db.select_all_stat()

    cnt1 = 0
    cnt2 = 0

    wb = Workbook()

    ws = wb.active
    ws["A1"] = "ID raqami"
    ws["B1"] = "ismi"
    ws["C1"] = "jami poralar soni"
    ws["D1"] = "qatnashgan xatmonalar soni"
    for n in all_stat:
        a = str(cnt1 + cnt2 + 2)
        ws["A" + a] = n[0]
        ws["B" + a] = n[1]
        ws["C" + a] = n[2]
        ws["D" + a] = n[3]
    wb.save("user_xatmonalar.xlsx")

    await msg.answer_document(types.InputFile(path_or_bytesio="user_xatmonalar.xlsx"),
                              caption="Фойдаланувчи жами неча пора олгани ва нечта хатмонада иштирок этганини кўрсатувчи жадвал")
    os.remove("user_xatmonalar.xlsx")
